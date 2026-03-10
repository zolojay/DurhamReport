"""
Phase 1 — Build Workbook Shell (Full Layout)
Creates the SCR Catalyst Test Report workbook with detailed layouts for every
visible and conditionally-visible sheet, plus backing infrastructure.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName

# ── Output path ──────────────────────────────────────────────────────────────
OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "output")
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "SCR_TestReport.xlsm")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# ── Color palette ────────────────────────────────────────────────────────────
COLORS = {
    "section_header":  "0E1638",
    "sub_section":     "2C3A8C",
    "column_header":   "4C5CE0",
    "input_cell":      "FFF9C4",
    "formula_output":  "E8E8E8",
    "label":           "E8EAF6",
    "pass":            "C8E6C9",
    "warning":         "FFE0B2",
    "fail":            "FFCDD2",
    "white":           "FFFFFF",
}

FILLS = {k: PatternFill(start_color=v, end_color=v, fill_type="solid")
         for k, v in COLORS.items()}

# ── Font definitions ─────────────────────────────────────────────────────────
FONT_TITLE_16     = Font(name="Calibri", size=16, bold=True, color=COLORS["white"])
FONT_SUBSEC_14    = Font(name="Calibri", size=14, bold=True, color=COLORS["white"])
FONT_SECTION_12   = Font(name="Calibri", size=12, bold=True, color=COLORS["white"])
FONT_BODY_11      = Font(name="Calibri", size=11)
FONT_BODY_10      = Font(name="Calibri", size=10)
FONT_BOLD_11      = Font(name="Calibri", size=11, bold=True)
FONT_COL_HDR      = Font(name="Calibri", size=11, bold=True, color=COLORS["white"])
FONT_BUTTON       = Font(name="Calibri", size=11, bold=True, color=COLORS["sub_section"])
FONT_ITALIC_11    = Font(name="Calibri", size=11, italic=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
BOTTOM_BORDER = Border(bottom=Side(style="thin"))

ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

# Row height constants
RH_SECTION = 24
RH_HEADER = 20
RH_DATA = 16

# ── Sheet definitions ────────────────────────────────────────────────────────
SHEET_DEFS = [
    ("Home",                "visible"),
    ("Specifications",      "visible"),
    ("HC Geometry",         "hidden"),
    ("Plate Geometry",      "hidden"),
    ("Corrugated Geometry", "hidden"),
    ("Setup Summary",       "hidden"),
    ("Activity",            "hidden"),
    ("Conversion",          "hidden"),
    ("DP",                  "hidden"),
    ("NOx Data",            "veryHidden"),
    ("NH3 Data",            "veryHidden"),
    ("SO2 Data",            "veryHidden"),
    ("SO3 Data",            "veryHidden"),
    ("Temperature Data",    "veryHidden"),
    ("Flow Data",           "veryHidden"),
    ("O2 Data",             "veryHidden"),
    ("DP Data",             "veryHidden"),
    ("Geometry Calc",       "veryHidden"),
    ("Control",             "veryHidden"),
    ("Product Specs",       "veryHidden"),
    ("Lists",               "veryHidden"),
    ("Constants",           "veryHidden"),
]

# ── Audit columns ────────────────────────────────────────────────────────────
AUDIT_COLUMNS = [
    "RecordID", "TestID", "EntryType", "DateEntered", "TimeEntered",
    "EnteredBy", "LastModifiedDate", "LastModifiedTime", "LastModifiedBy",
    "Excluded", "ExcludeReason", "Notes",
]

# ── Backing table definitions ────────────────────────────────────────────────
BACKING_TABLES = {
    "NOx Data": {
        "table_name": "tbl_NOx",
        "columns": AUDIT_COLUMNS + [
            "TestType", "SamplePoint", "Analyzer",
            "FTIR_NO_Wet", "FTIR_NO2_Wet", "FTIR_H2O_Pct",
            "NOxAn_NO_Dry", "NOxAn_NO2_Dry", "DryNOxUsed",
        ],
    },
    "NH3 Data": {
        "table_name": "tbl_NH3",
        "columns": AUDIT_COLUMNS + [
            "TestType", "SamplePoint", "Method",
            "FTIR_NH3_Wet", "FTIR_H2O_Pct",
            "IC_Result", "Dilution", "MeterVol_L", "MeterTemp_C", "BaroP_mmHg",
            "DryNH3",
        ],
    },
    "SO2 Data": {
        "table_name": "tbl_SO2",
        "columns": AUDIT_COLUMNS + [
            "TestType", "Stage", "Method",
            "FTIR_SO2_Wet", "FTIR_H2O_Pct",
            "IC_Result", "Dilution", "MeterVol_L", "MeterTemp_C", "BaroP_mmHg",
            "DrySO2",
        ],
    },
    "SO3 Data": {
        "table_name": "tbl_SO3",
        "columns": AUDIT_COLUMNS + [
            "TestType", "SamplePoint",
            "PullVol_L", "IC_Result", "Dilution", "MeterTemp_C", "RoomP_mmHg",
            "CorrectedGasVol", "DryMoles", "SO3Moles", "DrySO3",
        ],
    },
    "Temperature Data": {
        "table_name": "tbl_Temperature",
        "columns": AUDIT_COLUMNS + ["TestType", "Context", "Value_C"],
    },
    "Flow Data": {
        "table_name": "tbl_Flow",
        "columns": AUDIT_COLUMNS + ["TestType", "Context", "Value_scfm"],
    },
    "O2 Data": {
        "table_name": "tbl_O2",
        "columns": AUDIT_COLUMNS + ["TestType", "Context", "Value_Pct"],
    },
    "DP Data": {
        "table_name": "tbl_DP",
        "columns": AUDIT_COLUMNS + [
            "TestType",
            "DP_S2", "DP_S3", "DP_S4", "DP_S5",
            "DPTotal", "TheoryDP", "PctTheory", "Status",
        ],
    },
}

# ── Control sheet fields ─────────────────────────────────────────────────────
CONTROL_FIELDS = [
    ("Ctrl_TestID",                ""),
    ("Ctrl_WorkbookState",         "Setup"),
    ("Ctrl_ActiveGeometryType",    ""),
    ("Ctrl_TestStartTimestamp",    ""),
    ("Ctrl_SetupEnteredBy",        ""),
    ("Ctrl_SetupEnteredAt",        ""),
    ("Ctrl_Verified1By",           ""),
    ("Ctrl_Verified1At",           ""),
    ("Ctrl_Verified2By",           ""),
    ("Ctrl_Verified2At",           ""),
    ("Ctrl_SetupSignoffComplete",  "FALSE"),
]

# ── Constants ────────────────────────────────────────────────────────────────
CONSTANTS = [
    ("H2O_Reference",          0.18,    "H₂O reference for fallback chain (18%)"),
    ("MolarMass_NO",          30.01,    "g/mol"),
    ("MolarMass_NO2",         46.01,    "g/mol"),
    ("MolarMass_NH3",         17.03,    "g/mol"),
    ("MolarMass_SO2",         64.07,    "g/mol"),
    ("MolarMass_SO3",         80.06,    "g/mol"),
    ("STP_Temp_K",           273.15,    "Standard temperature (K)"),
    ("STP_Pressure_mmHg",    760.0,     "Standard pressure (mmHg)"),
    ("IdealGasVol_L",         22.414,   "Ideal gas molar volume at STP (L/mol)"),
    ("Nm3_to_scfm_Factor",    0.58858,  "Conversion factor Nm³/h → scfm"),
]

# ── Lists sheet content ──────────────────────────────────────────────────────
LISTS_DATA = {
    "SampleType":     ["Honeycomb", "Plate", "Corrugated"],
    "GeometryType":   ["HC", "Plate", "Corrugated"],
    "TestType":       ["Activity", "Conversion"],
    "SamplePoint":    ["Inlet", "Outlet"],
    "NOxAnalyzer":    ["FTIR", "NOxAn"],
    "NH3Method":      ["FTIR", "IC"],
    "SO2Method":      ["FTIR", "IC"],
    "SO2Stage":       ["Pre", "Test", "Post"],
    "TempContext":    ["Pre-Test", "During Test"],
    "FlowContext":    ["Pre-Test", "During Test"],
    "O2Context":      ["Pre-Test", "During Test"],
    "EntryType":      ["Form", "Manual", "Import"],
    "YesNo":          ["Yes", "No"],
    "TrueFalse":      ["TRUE", "FALSE"],
    "InletSO3Source": ["Average", "RecordID"],
    "SO2Source":      ["Validation", "Test Average", "RecordID"],
    "Technicians":    ["Tech1", "Tech2", "Tech3", "Tech4"],
}

# ── Specifications defaults ──────────────────────────────────────────────────
SPEC_FIELDS = [
    ("Spec_Temp_Tolerance_C",        5,     "Temperature tolerance (±°C)"),
    ("Spec_Flow_Tolerance_Pct",      5,     "Flow tolerance (±%)"),
    ("Spec_O2_Tolerance_Pct",        0.5,   "O₂ tolerance (±%)"),
    ("Spec_NOx_Tolerance_Pct",       5,     "NOx tolerance (±%)"),
    ("Spec_SO2_Tolerance_Pct",       10,    "SO₂ tolerance (±%)"),
    ("Spec_NH3_Tolerance_Pct",       10,    "NH₃ tolerance (±%)"),
    ("Spec_MR_Tolerance",            0.02,  "MR tolerance (±)"),
    ("Spec_SS_MinPoints",            4,     "Steady-state min points"),
    ("Spec_SS_K_StdDev_Max",         0.05,  "K StdDev threshold for steady-state"),
    ("Spec_SS_NormSlope_Max",        0.02,  "Normalized slope threshold"),
    ("Spec_SS_Conv_StdDev_Max",      2.0,   "Conversion StdDev threshold (%)"),
    ("Spec_SS_Conv_NormSlope_Max",   0.02,  "Conversion normalized slope threshold"),
    ("Spec_DP_PctTheory_Warning",    120,   "DP % Theory warning threshold"),
    ("Spec_DP_PctTheory_Fail",       150,   "DP % Theory fail threshold"),
]

# ── GC named ranges (placeholders for Phase 2) ──────────────────────────────
GC_NAMED_RANGES = [
    "GC_AvgAdjFFA",
    "GC_TotalAdjArea",
    "GC_ActiveLayers",
    "GC_Flow_Act_Nm3h",
    "GC_Flow_Conv_Nm3h",
    "GC_Flow_Act_scfm",
    "GC_Flow_Conv_scfm",
    "GC_AV_Act",
    "GC_AV_Conv",
    "GC_Flow_Act_Status",
    "GC_Flow_Conv_Status",
    "GC_SO3_Inj_Act",
    "GC_SO3_Inj_Conv",
    "GC_NH3_Inj_Act",
    "GC_NH3_Inj_Conv",
    "GC_SO2_Inj_Act",
    "GC_SO2_Inj_Conv",
]


# ═══════════════════════════════════════════════════════════════════════════════
# Helper functions
# ═══════════════════════════════════════════════════════════════════════════════

def apply_section_header(ws, row, col_start, col_end, text, level="section"):
    """Write a merged section-header band."""
    if level == "title":
        font = FONT_TITLE_16
        fill = FILLS["section_header"]
    elif level == "sub":
        font = FONT_SUBSEC_14
        fill = FILLS["sub_section"]
    else:
        font = FONT_SECTION_12
        fill = FILLS["section_header"]

    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = font
    cell.fill = fill
    cell.alignment = ALIGN_LEFT
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                        end_row=row, end_column=col_end)
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = fill
    ws.row_dimensions[row].height = RH_SECTION


def label_cell(ws, row, col, text):
    """Write a label-styled cell."""
    c = ws.cell(row=row, column=col, value=text)
    c.font = FONT_BOLD_11
    c.fill = FILLS["label"]
    c.alignment = ALIGN_RIGHT
    c.border = THIN_BORDER
    return c


def input_cell(ws, row, col, value=None, fmt=None):
    """Write an input-styled cell (#FFF9C4)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_BODY_11
    c.fill = FILLS["input_cell"]
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def output_cell(ws, row, col, value=None, fmt=None):
    """Write a formula-output-styled cell (#E8E8E8)."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = FONT_BODY_11
    c.fill = FILLS["formula_output"]
    c.alignment = ALIGN_CENTER
    c.border = THIN_BORDER
    if fmt:
        c.number_format = fmt
    return c


def col_headers(ws, row, col_start, headers):
    """Write column-header row (#4C5CE0 white bold)."""
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=col_start + i, value=h)
        c.font = FONT_COL_HDR
        c.fill = FILLS["column_header"]
        c.alignment = ALIGN_CENTER_WRAP
        c.border = THIN_BORDER
    ws.row_dimensions[row].height = RH_HEADER


def button_row(ws, row, col_start, col_end, text):
    """Write a button placeholder row — merged, styled."""
    cell = ws.cell(row=row, column=col_start, value=text)
    cell.font = FONT_BUTTON
    cell.fill = FILLS["label"]
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    if col_end > col_start:
        ws.merge_cells(start_row=row, start_column=col_start,
                        end_row=row, end_column=col_end)
    for c in range(col_start + 1, col_end + 1):
        ws.cell(row=row, column=c).fill = FILLS["label"]
        ws.cell(row=row, column=c).border = THIN_BORDER


def set_col_widths(ws, widths):
    """Set column widths from dict {col_letter: width}."""
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def data_rows_fill(ws, row_start, row_end, col_start, col_end, fill_key="formula_output"):
    """Fill a block of data rows with a given style."""
    fill = FILLS[fill_key]
    for r in range(row_start, row_end + 1):
        ws.row_dimensions[r].height = RH_DATA
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY_11
            cell.fill = fill
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER


def create_excel_table(ws, table_name, headers, header_row, col_start, data_rows=1):
    """Create an Excel Table with headers and empty data rows."""
    col_end = col_start + len(headers) - 1
    end_row = header_row + data_rows
    col_headers(ws, header_row, col_start, headers)
    for r in range(header_row + 1, end_row + 1):
        for c in range(col_start, col_end + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY_10
            cell.border = THIN_BORDER
    ref = f"{get_column_letter(col_start)}{header_row}:{get_column_letter(col_end)}{end_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleLight1", showFirstColumn=False,
        showLastColumn=False, showRowStripes=False, showColumnStripes=False,
    )
    ws.add_table(table)
    return table


def add_dv_list(ws, named_range, cells_range):
    """Add data-validation dropdown referencing a named range."""
    dv = DataValidation(type="list", formula1=f"={named_range}", allow_blank=True)
    dv.sqref = cells_range
    ws.add_data_validation(dv)


def hide_gridlines(ws):
    """Hide gridlines on sheet."""
    ws.sheet_view.showGridLines = False


def freeze_at(ws, cell_ref):
    """Freeze panes at given cell."""
    ws.freeze_panes = cell_ref


# ═══════════════════════════════════════════════════════════════════════════════
# 1. HOME SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_home(ws):
    set_col_widths(ws, {"A": 2, "B": 22, "C": 18, "D": 18, "E": 18})
    hide_gridlines(ws)

    # ── Title row 1 ──────────────────────────────────────────────────────
    apply_section_header(ws, 1, 2, 4, "SCR Catalyst Test Report", level="title")
    ws.row_dimensions[1].height = 30

    # ── Identity block rows 3-11 ─────────────────────────────────────────
    apply_section_header(ws, 3, 2, 4, "Test Identity", level="sub")

    identity_fields = [
        "LRF #", "Load ID", "Project Name", "Date",
        "Sample Type", "Active Technician", "SO₂ Gas On", "NH₃ Gas On",
    ]
    for i, lbl in enumerate(identity_fields):
        r = 4 + i
        label_cell(ws, r, 2, lbl)
        input_cell(ws, r, 3)
        ws.row_dimensions[r].height = RH_DATA

    # Data validations for dropdowns
    add_dv_list(ws, "List_SampleType", "C8")    # Sample Type row 8
    add_dv_list(ws, "List_Technicians", "C9")    # Active Technician row 9
    # Time format for SO2/NH3 Gas On
    ws["C10"].number_format = "h:mm"
    ws["C11"].number_format = "h:mm"

    # ── Test Conditions block rows 13-25 ─────────────────────────────────
    apply_section_header(ws, 13, 2, 4, "Test Conditions", level="sub")
    col_headers(ws, 14, 2, ["Variable", "Activity", "Conversion"])

    condition_vars = [
        "AV", "UGS", "Temperature (°C)", "H₂O (%)", "O₂ (%)",
        "SO₂ (ppmvd)", "SO₃ (ppmvd)", "NOx (ppmvd)", "MR",
    ]
    for i, var in enumerate(condition_vars):
        r = 15 + i
        label_cell(ws, r, 2, var)
        input_cell(ws, r, 3)  # Activity
        input_cell(ws, r, 4)  # Conversion
        ws.row_dimensions[r].height = RH_DATA

    # Flow Source/Status — formula output row 24
    label_cell(ws, 24, 2, "Flow Source / Status")
    output_cell(ws, 24, 3, "=GC_Flow_Act_Status")
    output_cell(ws, 24, 4, "=GC_Flow_Conv_Status")
    ws.row_dimensions[24].height = RH_DATA

    # Flow (Nm³/h) — formula output row 25
    label_cell(ws, 25, 2, "Flow (Nm³/h)")
    output_cell(ws, 25, 3, "=GC_Flow_Act_Nm3h")
    output_cell(ws, 25, 4, "=GC_Flow_Conv_Nm3h")
    ws.row_dimensions[25].height = RH_DATA

    # ── Workflow Buttons block rows 28-33 ────────────────────────────────
    apply_section_header(ws, 28, 2, 4, "Workflow", level="sub")
    buttons = [
        "[ Set Up Geometry ]", "[ Open Setup Summary ]",
        "[ Verify Setup ]", "[ Begin Test ]", "[ New Test ]",
    ]
    for i, btn in enumerate(buttons):
        r = 29 + i
        button_row(ws, r, 2, 4, btn)
        ws.row_dimensions[r].height = RH_DATA

    # ── Readiness Status block rows 36-46 ────────────────────────────────
    apply_section_header(ws, 36, 2, 3, "Readiness Status", level="sub")
    status_items = [
        "Sample type selected", "Geometry entered", "Geometry resolved",
        "Flow resolved", "Primary verification complete",
        "Secondary verification complete", "Ready to begin test",
        "Activity data present", "Conversion data present", "DP data present",
    ]
    for i, item in enumerate(status_items):
        r = 37 + i
        label_cell(ws, r, 2, item)
        output_cell(ws, r, 3, "\u2014")  # em-dash
        ws.row_dimensions[r].height = RH_DATA

    ws.sheet_properties.tabColor = "0E1638"
    freeze_at(ws, "B2")
    ws.print_area = "A1:E46"


# ═══════════════════════════════════════════════════════════════════════════════
# 2. SPECIFICATIONS SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_specifications(ws):
    set_col_widths(ws, {"A": 2, "B": 30, "C": 15, "D": 40})
    hide_gridlines(ws)

    apply_section_header(ws, 1, 2, 4, "Specifications & Thresholds", level="title")
    col_headers(ws, 3, 2, ["Parameter", "Value", "Description"])

    for i, (name, value, desc) in enumerate(SPEC_FIELDS):
        r = 4 + i
        label_cell(ws, r, 2, name)
        input_cell(ws, r, 3, value)
        lbl = ws.cell(row=r, column=4, value=desc)
        lbl.font = FONT_BODY_10
        lbl.fill = FILLS["label"]
        lbl.border = THIN_BORDER
        lbl.alignment = ALIGN_LEFT
        ws.row_dimensions[r].height = RH_DATA

    freeze_at(ws, "B4")
    ws.print_area = f"A1:D{3 + len(SPEC_FIELDS)}"


# ═══════════════════════════════════════════════════════════════════════════════
# 3. HC GEOMETRY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_hc_geometry(ws):
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 12, "D": 12, "E": 12,
        "F": 12, "G": 12, "H": 12, "I": 12, "J": 14, "K": 14,
    })
    hide_gridlines(ws)

    # Header
    apply_section_header(ws, 1, 2, 11, "Honeycomb Geometry", level="title")
    # Instruction row
    instr = ws.cell(row=2, column=2, value="Enter measurements for each active layer")
    instr.font = FONT_ITALIC_11
    instr.fill = FILLS["label"]
    instr.alignment = ALIGN_LEFT
    ws.merge_cells("B2:K2")
    for c in range(3, 12):
        ws.cell(row=2, column=c).fill = FILLS["label"]

    # Column headers row 4
    headers = [
        "Layer", "Product Type", "AP Override", "Length (mm)",
        "Cells A", "Cells B", "Width A (mm)", "Width B (mm)",
        "Plugged Cells", "Status",
    ]
    col_headers(ws, 4, 2, headers)
    ws.row_dimensions[4].height = 24

    # Layer rows 5-10
    for layer in range(1, 7):
        r = 4 + layer
        # Layer number
        lc = ws.cell(row=r, column=2, value=layer)
        lc.font = FONT_BOLD_11
        lc.fill = FILLS["label"]
        lc.alignment = ALIGN_CENTER
        lc.border = THIN_BORDER
        # Product Type (col C=3)
        input_cell(ws, r, 3)
        # AP Override (col D=4)
        input_cell(ws, r, 4)
        # Length, Cells A, Cells B, Width A, Width B, Plugged (cols 5-10)
        for c in range(5, 11):
            input_cell(ws, r, c)
        # Status (col K=11)
        output_cell(ws, r, 11)
        ws.row_dimensions[r].height = RH_DATA

    # Convenience Outputs
    apply_section_header(ws, 13, 2, 5, "Resolved Outputs (read-only)", level="sub")
    outputs = [
        "Total Active Layers", "Avg Adjusted FFA (m²)",
        "Total Adjusted Area (m²)", "Pitch (mm)", "Hydraulic Diameter (mm)",
    ]
    for i, lbl in enumerate(outputs):
        r = 14 + i
        label_cell(ws, r, 2, lbl)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c2 in range(2, 4):
            ws.cell(row=r, column=c2).fill = FILLS["label"]
            ws.cell(row=r, column=c2).border = THIN_BORDER
        output_cell(ws, r, 4)
        ws.row_dimensions[r].height = RH_DATA

    # Navigation
    button_row(ws, 20, 2, 4, "[ Back to Home ]")

    freeze_at(ws, "B5")
    ws.print_area = "A1:K20"


# ═══════════════════════════════════════════════════════════════════════════════
# 4. PLATE GEOMETRY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_plate_geometry(ws):
    set_col_widths(ws, {"A": 2, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14})
    hide_gridlines(ws)

    apply_section_header(ws, 1, 2, 6, "Plate Geometry", level="title")

    # Total Plates
    label_cell(ws, 3, 2, "Total Plates")
    input_cell(ws, 3, 3)

    # ── Box 1 ────────────────────────────────────────────────────────────
    apply_section_header(ws, 5, 2, 6, "Box 1 Measurements", level="sub")
    box1_hdrs = ["Plate #", "Length (mm)", "Thickness (mm)", "Width A (mm)", "Width B (mm)"]
    col_headers(ws, 6, 2, box1_hdrs)

    for plate in range(1, 14):
        r = 6 + plate
        lc = ws.cell(row=r, column=2, value=plate)
        lc.font = FONT_BOLD_11
        lc.fill = FILLS["label"]
        lc.alignment = ALIGN_CENTER
        lc.border = THIN_BORDER
        for c in range(3, 7):
            input_cell(ws, r, c, fmt="0.00")
        ws.row_dimensions[r].height = RH_DATA

    # Box 1 Average row 20
    r_avg1 = 20
    lbl = ws.cell(row=r_avg1, column=2, value="Average")
    lbl.font = FONT_BOLD_11
    lbl.fill = FILLS["label"]
    lbl.alignment = ALIGN_CENTER
    lbl.border = THIN_BORDER
    for c in range(3, 7):
        col_l = get_column_letter(c)
        output_cell(ws, r_avg1, c, f"=AVERAGE({col_l}7:{col_l}19)", fmt="0.00")

    # ── Box 2 ────────────────────────────────────────────────────────────
    apply_section_header(ws, 22, 2, 6, "Box 2 Measurements", level="sub")
    col_headers(ws, 23, 2, box1_hdrs)

    for plate in range(1, 14):
        r = 23 + plate
        lc = ws.cell(row=r, column=2, value=plate)
        lc.font = FONT_BOLD_11
        lc.fill = FILLS["label"]
        lc.alignment = ALIGN_CENTER
        lc.border = THIN_BORDER
        for c in range(3, 7):
            input_cell(ws, r, c, fmt="0.00")
        ws.row_dimensions[r].height = RH_DATA

    # Box 2 Average row 37
    r_avg2 = 37
    lbl = ws.cell(row=r_avg2, column=2, value="Average")
    lbl.font = FONT_BOLD_11
    lbl.fill = FILLS["label"]
    lbl.alignment = ALIGN_CENTER
    lbl.border = THIN_BORDER
    for c in range(3, 7):
        col_l = get_column_letter(c)
        output_cell(ws, r_avg2, c, f"=AVERAGE({col_l}24:{col_l}36)", fmt="0.00")

    # ── Combined Summary ─────────────────────────────────────────────────
    apply_section_header(ws, 39, 2, 5, "Combined Summary (read-only)", level="sub")
    summary = [
        ("Overall Average Length", f"=AVERAGE(C20,C37)"),
        ("Overall Average Thickness", f"=AVERAGE(D20,D37)"),
        ("Overall Average Width A", f"=AVERAGE(E20,E37)"),
        ("Overall Average Width B", f"=AVERAGE(F20,F37)"),
    ]
    for i, (lbl_text, formula) in enumerate(summary):
        r = 40 + i
        label_cell(ws, r, 2, lbl_text)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c2 in range(2, 4):
            ws.cell(row=r, column=c2).fill = FILLS["label"]
            ws.cell(row=r, column=c2).border = THIN_BORDER
        output_cell(ws, r, 4, formula, fmt="0.00")
        ws.row_dimensions[r].height = RH_DATA

    # Navigation
    button_row(ws, 45, 2, 4, "[ Back to Home ]")

    freeze_at(ws, "B4")
    ws.print_area = "A1:F45"


# ═══════════════════════════════════════════════════════════════════════════════
# 5. CORRUGATED GEOMETRY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_corrugated_geometry(ws):
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14,
    })
    hide_gridlines(ws)

    apply_section_header(ws, 1, 2, 8, "Corrugated Geometry", level="title")

    # Column headers row 3
    headers = [
        "Layer", "SSA (m²/m³)", "Length (mm)", "Width (mm)",
        "Height (mm)", "Total Cells", "Plugged Cells",
    ]
    col_headers(ws, 3, 2, headers)

    for layer in range(1, 7):
        r = 3 + layer
        lc = ws.cell(row=r, column=2, value=layer)
        lc.font = FONT_BOLD_11
        lc.fill = FILLS["label"]
        lc.alignment = ALIGN_CENTER
        lc.border = THIN_BORDER
        for c in range(3, 9):
            input_cell(ws, r, c)
        ws.row_dimensions[r].height = RH_DATA

    # Resolved Outputs
    apply_section_header(ws, 11, 2, 5, "Resolved Outputs (read-only)", level="sub")
    outputs = [
        "Total Active Layers", "Avg Adjusted FFA (m²)",
        "Total Adjusted Area (m²)", "Active Volume (m³)",
    ]
    for i, lbl_text in enumerate(outputs):
        r = 12 + i
        label_cell(ws, r, 2, lbl_text)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c2 in range(2, 4):
            ws.cell(row=r, column=c2).fill = FILLS["label"]
            ws.cell(row=r, column=c2).border = THIN_BORDER
        output_cell(ws, r, 4)
        ws.row_dimensions[r].height = RH_DATA

    # Navigation
    button_row(ws, 17, 2, 4, "[ Back to Home ]")

    freeze_at(ws, "B4")
    ws.print_area = "A1:H17"


# ═══════════════════════════════════════════════════════════════════════════════
# 6. SETUP SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_setup_summary(ws):
    set_col_widths(ws, {"A": 2, "B": 28, "C": 20, "D": 20, "E": 20})
    hide_gridlines(ws)

    apply_section_header(ws, 1, 2, 5, "Setup Summary", level="title")

    # ── Geometry & Flow Results rows 3-14 ────────────────────────────────
    apply_section_header(ws, 3, 2, 5, "Geometry & Flow Results", level="sub")
    col_headers(ws, 4, 2, ["Parameter", "Activity", "Conversion"])

    # Rows 5-7: single-value geometry outputs (no Act/Conv split)
    geo_single = [
        ("Avg Adjusted FFA (m²)", "=GC_AvgAdjFFA"),
        ("Total Adjusted Area (m²)", "=GC_TotalAdjArea"),
        ("Active Layers", "=GC_ActiveLayers"),
    ]
    for i, (lbl_text, formula) in enumerate(geo_single):
        r = 5 + i
        label_cell(ws, r, 2, lbl_text)
        output_cell(ws, r, 3, formula)
        output_cell(ws, r, 4)  # blank for single-value rows
        ws.row_dimensions[r].height = RH_DATA

    # Rows 8-14: Activity/Conversion split
    flow_rows = [
        ("Flow (Nm³/h)",        "=GC_Flow_Act_Nm3h",    "=GC_Flow_Conv_Nm3h"),
        ("Flow (scfm)",         "=GC_Flow_Act_scfm",    "=GC_Flow_Conv_scfm"),
        ("AV Used",             "=GC_AV_Act",           "=GC_AV_Conv"),
        ("Flow Source / Status","=GC_Flow_Act_Status",  "=GC_Flow_Conv_Status"),
    ]
    for i, (lbl_text, f_act, f_conv) in enumerate(flow_rows):
        r = 8 + i
        label_cell(ws, r, 2, lbl_text)
        output_cell(ws, r, 3, f_act)
        output_cell(ws, r, 4, f_conv)
        ws.row_dimensions[r].height = RH_DATA

    # ── Injection Rates rows 16-22 ───────────────────────────────────────
    apply_section_header(ws, 16, 2, 4, "Injection Rates", level="sub")
    col_headers(ws, 17, 2, ["Parameter", "Activity", "Conversion"])

    inj_rows = [
        ("SO₃ Injection (mL/min)", "=GC_SO3_Inj_Act",  "=GC_SO3_Inj_Conv"),
        ("NH₃ Injection (mL/min)", "=GC_NH3_Inj_Act",  "=GC_NH3_Inj_Conv"),
        ("SO₂ Injection (mL/min)", "=GC_SO2_Inj_Act",  "=GC_SO2_Inj_Conv"),
        ("Combustion NH₃ Est.",     None,                None),
    ]
    for i, (lbl_text, f_act, f_conv) in enumerate(inj_rows):
        r = 18 + i
        label_cell(ws, r, 2, lbl_text)
        output_cell(ws, r, 3, f_act)
        output_cell(ws, r, 4, f_conv)
        ws.row_dimensions[r].height = RH_DATA

    # ── Slip Prediction rows 25-30 ───────────────────────────────────────
    apply_section_header(ws, 25, 2, 4, "Slip Prediction", level="sub")
    label_cell(ws, 26, 2, "Expected K")
    input_cell(ws, 26, 3)  # editable
    ws.row_dimensions[26].height = RH_DATA

    slip_outputs = [
        "Predicted Slip (ppmvd)", "Predicted Outlet NOx",
        "Predicted DeNOx (%)",
    ]
    for i, lbl_text in enumerate(slip_outputs):
        r = 27 + i
        label_cell(ws, r, 2, lbl_text)
        output_cell(ws, r, 3)
        ws.row_dimensions[r].height = RH_DATA

    # ── Setup Verification rows 33-41 ────────────────────────────────────
    apply_section_header(ws, 33, 2, 4, "Setup Verification", level="sub")
    verify_rows = [
        ("Setup Entered By",   "=Ctrl_SetupEnteredBy"),
        ("Setup Entered At",   "=Ctrl_SetupEnteredAt"),
        ("Verification 1 By",  "=Ctrl_Verified1By"),
        ("Verification 1 At",  "=Ctrl_Verified1At"),
        ("Verification 2 By",  "=Ctrl_Verified2By"),
        ("Verification 2 At",  "=Ctrl_Verified2At"),
        ("Signoff Status",     "=Ctrl_SetupSignoffComplete"),
    ]
    for i, (lbl_text, formula) in enumerate(verify_rows):
        r = 34 + i
        label_cell(ws, r, 2, lbl_text)
        output_cell(ws, r, 3, formula)
        ws.row_dimensions[r].height = RH_DATA

    # Navigation buttons
    button_row(ws, 43, 2, 4, "[ Verify Setup ]")
    button_row(ws, 44, 2, 4, "[ Back to Home ]")

    freeze_at(ws, "B2")
    ws.print_area = "A1:E44"


# ═══════════════════════════════════════════════════════════════════════════════
# 7. ACTIVITY DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def build_activity(ws):
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 12, "D": 12, "E": 12, "F": 12,
        "G": 12, "H": 10, "I": 12, "J": 12, "K": 12, "L": 12,
        "M": 12, "N": 12, "O": 12, "P": 14,
    })
    hide_gridlines(ws)

    # ── Title row 1 ──────────────────────────────────────────────────────
    apply_section_header(ws, 1, 2, 16, "Activity Dashboard", level="title")

    # ── Inline banner row 3 ──────────────────────────────────────────────
    banner = [
        (2, "Test ID:", True),  (3, "=Ctrl_TestID", False),
        (4, "Sample:", True),   (5, None, False),
        (6, "Flow:", True),     (7, "=GC_Flow_Act_Nm3h", False),
        (8, "Check 1:", True),  (9, "=Ctrl_Verified1By", False),
        (10, "Check 2:", True), (11, "=Ctrl_Verified2By", False),
    ]
    for col, val, is_label in banner:
        c = ws.cell(row=3, column=col, value=val)
        c.font = FONT_BOLD_11 if is_label else FONT_BODY_11
        c.fill = FILLS["label"] if is_label else FILLS["formula_output"]
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER
    ws.row_dimensions[3].height = RH_DATA

    # ── Pre-Test Validation rows 6-16 ────────────────────────────────────
    apply_section_header(ws, 6, 2, 8, "Pre-Test Validation", level="sub")
    col_headers(ws, 7, 2, ["Parameter", "Actual", "Target", "LSL", "USL", "Status"])

    val_params = [
        "Temperature", "Flow", "O₂", "NOx", "SO₂",
        "NH₃ (MR>0)", "MR (MR>0)",
    ]
    for i, param in enumerate(val_params):
        r = 8 + i
        label_cell(ws, r, 2, param)
        for c in range(3, 7):
            output_cell(ws, r, c)
        output_cell(ws, r, 7)  # Status — will get conditional formatting
        ws.row_dimensions[r].height = RH_DATA

    # Row 16: form launch buttons
    btn_defs = [(2, 3, "[ Log Physical ]"), (4, 5, "[ Log NOx ]"),
                (6, 7, "[ Log SO2 ]"), (8, 9, "[ Log NH3 ]")]
    for cs, ce, txt in btn_defs:
        button_row(ws, 16, cs, ce, txt)

    # ── Activity Results rows 18-25 ──────────────────────────────────────
    apply_section_header(ws, 18, 2, 16, "Activity Results", level="sub")
    result_hdrs = [
        "Use", "Inlet NOx ID", "Inlet NOx", "Outlet NOx ID", "Outlet NOx",
        "DeNOx %", "K", "NH3 ID", "NH3 Used", "NH3 Pt",
        "MR Actual", "Corm-K", "H2O Used", "H2O K", "Browse",
    ]
    col_headers(ws, 19, 2, result_hdrs)
    ws.row_dimensions[19].height = 30

    for pass_i in range(6):
        r = 20 + pass_i
        # Use dropdown
        input_cell(ws, r, 2)
        # Inlet NOx ID
        input_cell(ws, r, 3)
        # Inlet NOx value
        output_cell(ws, r, 4)
        # Outlet NOx ID
        input_cell(ws, r, 5)
        # Outlet NOx value
        output_cell(ws, r, 6)
        # DeNOx %
        output_cell(ws, r, 7)
        # K
        output_cell(ws, r, 8)
        # NH3 ID
        input_cell(ws, r, 9)
        # NH3 Used
        output_cell(ws, r, 10)
        # NH3 Pt
        output_cell(ws, r, 11)
        # MR Actual
        output_cell(ws, r, 12)
        # Corm-K
        output_cell(ws, r, 13)
        # H2O Used
        output_cell(ws, r, 14)
        # H2O K
        output_cell(ws, r, 15)
        # Browse placeholder
        c = ws.cell(row=r, column=16, value="Browse")
        c.font = FONT_BUTTON
        c.fill = FILLS["label"]
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.row_dimensions[r].height = RH_DATA

    # Data validation for Use column
    add_dv_list(ws, "List_YesNo", "B20:B25")

    # ── Steady-State Check rows 28-33 ────────────────────────────────────
    apply_section_header(ws, 28, 2, 8, "Steady-State Check", level="sub")
    ss_items = ["K Mean", "K StdDev", "Normalized Slope",
                "Steady-State Result", "Points Used"]
    for i, item in enumerate(ss_items):
        r = 29 + i
        label_cell(ws, r, 2, item)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c2 in [2, 3]:
            ws.cell(row=r, column=c2).fill = FILLS["label"]
            ws.cell(row=r, column=c2).border = THIN_BORDER
        output_cell(ws, r, 4)
        ws.row_dimensions[r].height = RH_DATA

    # ── Compact Reading Tables ───────────────────────────────────────────
    _build_compact_table(ws, 35, "Inlet NOx Readings",
        ["RecordID", "NOx (ppmvd)", "Analyzer", "H2O (%)", "Time", "Technician", "Excluded"])
    _build_compact_table(ws, 45, "Outlet NOx Readings",
        ["RecordID", "NOx (ppmvd)", "Analyzer", "H2O (%)", "Time", "Technician", "Excluded"])
    _build_compact_table(ws, 55, "SO₂ Readings",
        ["RecordID", "SO2 (ppmvd)", "Stage", "Method", "H2O (%)", "Time", "Technician"])
    _build_compact_table(ws, 65, "NH₃ Readings",
        ["RecordID", "NH3 (ppmvd)", "Point", "Method", "H2O (%)", "Time", "Technician"])

    freeze_at(ws, "B2")
    ws.print_area = "A1:P73"


# ═══════════════════════════════════════════════════════════════════════════════
# 8. CONVERSION DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def build_conversion(ws):
    set_col_widths(ws, {
        "A": 2, "B": 14, "C": 12, "D": 14, "E": 12, "F": 14,
        "G": 12, "H": 12, "I": 12, "J": 12, "K": 12, "L": 14,
    })
    hide_gridlines(ws)

    # ── Title ────────────────────────────────────────────────────────────
    apply_section_header(ws, 1, 2, 12, "Conversion Dashboard", level="title")

    # ── Inline banner row 3 ──────────────────────────────────────────────
    banner = [
        (2, "Test ID:", True),  (3, "=Ctrl_TestID", False),
        (4, "Sample:", True),   (5, None, False),
        (6, "Flow:", True),     (7, "=GC_Flow_Conv_Nm3h", False),
        (8, "Check 1:", True),  (9, "=Ctrl_Verified1By", False),
        (10, "Check 2:", True), (11, "=Ctrl_Verified2By", False),
    ]
    for col, val, is_label in banner:
        c = ws.cell(row=3, column=col, value=val)
        c.font = FONT_BOLD_11 if is_label else FONT_BODY_11
        c.fill = FILLS["label"] if is_label else FILLS["formula_output"]
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER
    ws.row_dimensions[3].height = RH_DATA

    # ── Pre-Test Validation rows 6-17 ────────────────────────────────────
    apply_section_header(ws, 6, 2, 8, "Pre-Test Validation", level="sub")
    col_headers(ws, 7, 2, ["Parameter", "Actual", "Target", "LSL", "USL", "Status"])

    val_params = [
        "Temperature", "Flow", "O₂", "NOx", "SO₂", "SO₃",
        "NH₃ (MR>0)", "MR (MR>0)",
    ]
    for i, param in enumerate(val_params):
        r = 8 + i
        label_cell(ws, r, 2, param)
        for c in range(3, 7):
            output_cell(ws, r, c)
        output_cell(ws, r, 7)  # Status
        ws.row_dimensions[r].height = RH_DATA

    # Row 17: form launch buttons
    btn_defs = [(2, 3, "[ Log Physical ]"), (4, 5, "[ Log SO3 ]"),
                (6, 7, "[ Log SO2 ]"), (8, 9, "[ Log NOx ]"),
                (10, 11, "[ Log NH3 ]")]
    for cs, ce, txt in btn_defs:
        button_row(ws, 17, cs, ce, txt)

    # ── Conversion Results rows 19-26 ────────────────────────────────────
    apply_section_header(ws, 19, 2, 12, "Conversion Results", level="sub")
    result_hdrs = [
        "Use", "Out SO3 ID", "Out SO3", "In SO3 Src", "In SO3",
        "SO2 Src", "SO2 Used", "Difference", "Conv %", "NH3 ID", "Browse",
    ]
    col_headers(ws, 20, 2, result_hdrs)
    ws.row_dimensions[20].height = 30

    for pass_i in range(6):
        r = 21 + pass_i
        input_cell(ws, r, 2)   # Use
        input_cell(ws, r, 3)   # Outlet SO3 RecordID
        output_cell(ws, r, 4)  # Outlet SO3 value
        input_cell(ws, r, 5)   # Inlet SO3 Source
        output_cell(ws, r, 6)  # Inlet SO3 value
        input_cell(ws, r, 7)   # SO2 Source
        output_cell(ws, r, 8)  # SO2 Used
        output_cell(ws, r, 9)  # Difference
        output_cell(ws, r, 10) # Conv %
        input_cell(ws, r, 11)  # NH3 RecordID
        c = ws.cell(row=r, column=12, value="Browse")
        c.font = FONT_BUTTON
        c.fill = FILLS["label"]
        c.alignment = ALIGN_CENTER
        c.border = THIN_BORDER
        ws.row_dimensions[r].height = RH_DATA

    # Data validations
    add_dv_list(ws, "List_YesNo", "B21:B26")
    add_dv_list(ws, "List_InletSO3Source", "E21:E26")
    add_dv_list(ws, "List_SO2Source", "G21:G26")

    # ── Steady-State Check rows 29-34 ────────────────────────────────────
    apply_section_header(ws, 29, 2, 8, "Steady-State Check", level="sub")
    ss_items = ["Conv % Mean", "Conv % StdDev", "Normalized Slope",
                "Steady-State Result", "Points Used"]
    for i, item in enumerate(ss_items):
        r = 30 + i
        label_cell(ws, r, 2, item)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        for c2 in [2, 3]:
            ws.cell(row=r, column=c2).fill = FILLS["label"]
            ws.cell(row=r, column=c2).border = THIN_BORDER
        output_cell(ws, r, 4)
        ws.row_dimensions[r].height = RH_DATA

    # ── Compact Reading Tables ───────────────────────────────────────────
    _build_compact_table(ws, 36, "SO₃ Outlet Readings",
        ["RecordID", "SO3 (ppmvd)", "Pull Vol", "Time", "Technician", "Excluded"])
    _build_compact_table(ws, 46, "SO₃ Inlet Readings",
        ["RecordID", "SO3 (ppmvd)", "Pull Vol", "Time", "Technician", "Excluded"])
    _build_compact_table(ws, 56, "SO₂ Readings",
        ["RecordID", "SO2 (ppmvd)", "Stage", "Method", "Time", "Technician"])
    _build_compact_table(ws, 66, "NOx Readings",
        ["RecordID", "NOx (ppmvd)", "Analyzer", "Time", "Technician"])
    _build_compact_table(ws, 76, "NH₃ Readings",
        ["RecordID", "NH3 (ppmvd)", "Point", "Method", "H2O (%)", "Time", "Technician"])

    freeze_at(ws, "B2")
    ws.print_area = "A1:L84"


# ═══════════════════════════════════════════════════════════════════════════════
# Compact reading table helper (used by Activity & Conversion dashboards)
# ═══════════════════════════════════════════════════════════════════════════════

def _build_compact_table(ws, start_row, title, headers):
    """Build a compact reading table with header, 6 display rows, and buttons."""
    end_col = 2 + len(headers) - 1
    apply_section_header(ws, start_row, 2, end_col, title, level="section")
    col_headers(ws, start_row + 1, 2, headers)
    # 6 display rows
    data_rows_fill(ws, start_row + 2, start_row + 7, 2, end_col, "formula_output")
    # Button row
    btn_r = start_row + 8
    button_row(ws, btn_r, 2, 3, "[ + Add ]")
    button_row(ws, btn_r, 4, 5, "[ Browse ]")


# ═══════════════════════════════════════════════════════════════════════════════
# 9. DP DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════

def build_dp(ws):
    set_col_widths(ws, {
        "A": 2, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12,
        "G": 12, "H": 14, "I": 14, "J": 14, "K": 14,
    })
    hide_gridlines(ws)

    apply_section_header(ws, 1, 2, 11, "Differential Pressure", level="title")

    # Inline banner row 3
    banner = [
        (2, "Test ID:", True),  (3, "=Ctrl_TestID", False),
        (4, "Check 1:", True),  (5, "=Ctrl_Verified1By", False),
        (6, "Check 2:", True),  (7, "=Ctrl_Verified2By", False),
    ]
    for col, val, is_label in banner:
        c = ws.cell(row=3, column=col, value=val)
        c.font = FONT_BOLD_11 if is_label else FONT_BODY_11
        c.fill = FILLS["label"] if is_label else FILLS["formula_output"]
        c.border = THIN_BORDER
        c.alignment = ALIGN_CENTER
    ws.row_dimensions[3].height = RH_DATA

    # DP Measurements rows 5-12
    apply_section_header(ws, 5, 2, 11, "DP Measurements", level="sub")
    dp_hdrs = [
        "RecordID", "Test Type", "DP@S2", "DP@S3", "DP@S4",
        "DP@S5", "DP Total", "Theory DP", "% Theory", "Status",
    ]
    col_headers(ws, 6, 2, dp_hdrs)

    # 6 display rows
    data_rows_fill(ws, 7, 12, 2, 11, "formula_output")
    # Status column (K=11) — will get conditional formatting in Phase 7
    # For now just mark it as output

    # Buttons
    button_row(ws, 14, 2, 4, "[ Log DP ]")
    button_row(ws, 16, 2, 4, "[ Back to Home ]")

    freeze_at(ws, "B2")
    ws.print_area = "A1:K16"


# ═══════════════════════════════════════════════════════════════════════════════
# 10. GEOMETRY CALC (VeryHidden)
# ═══════════════════════════════════════════════════════════════════════════════

def build_geometry_calc(ws):
    set_col_widths(ws, {"A": 3, "B": 30, "C": 20, "D": 20, "E": 3})

    apply_section_header(ws, 1, 2, 4, "Geometry Calc Engine", level="title")

    sections = [
        (3,  "Active Geometry / Mode Selection"),
        (8,  "Unified Geometry Outputs"),
        (16, "Honeycomb Calculations"),
        (24, "Plate Calculations"),
        (32, "Corrugated Calculations"),
        (40, "Flow Calculations"),
        (50, "Injection Rate Calculations"),
        (60, "Shared Resolved Outputs"),
    ]
    for r, title in sections:
        apply_section_header(ws, r, 2, 4, title, level="section")

    # GC_ output cells
    for i, name in enumerate(GC_NAMED_RANGES):
        r = 61 + i
        label_cell(ws, r, 2, name)
        output_cell(ws, r, 3)
        ws.row_dimensions[r].height = RH_DATA


# ═══════════════════════════════════════════════════════════════════════════════
# Infrastructure sheets (Control, Constants, Lists, Product Specs, Backing)
# ═══════════════════════════════════════════════════════════════════════════════

def build_control(ws):
    set_col_widths(ws, {"A": 3, "B": 35, "C": 30, "D": 3})
    apply_section_header(ws, 1, 2, 3, "Workbook Control", level="title")
    for i, (field, default) in enumerate(CONTROL_FIELDS):
        r = 3 + i
        label_cell(ws, r, 2, field)
        output_cell(ws, r, 3, default)
        ws.row_dimensions[r].height = RH_DATA


def build_constants(ws):
    set_col_widths(ws, {"A": 3, "B": 30, "C": 18, "D": 40, "E": 3})
    apply_section_header(ws, 1, 2, 4, "Constants", level="title")
    col_headers(ws, 3, 2, ["Name", "Value", "Description"])
    for i, (name, value, desc) in enumerate(CONSTANTS):
        r = 4 + i
        label_cell(ws, r, 2, name)
        output_cell(ws, r, 3, value)
        lbl = ws.cell(row=r, column=4, value=desc)
        lbl.font = FONT_BODY_10
        lbl.fill = FILLS["label"]
        lbl.border = THIN_BORDER
        ws.row_dimensions[r].height = RH_DATA


def build_lists(ws):
    set_col_widths(ws, {"A": 3})
    apply_section_header(ws, 1, 2, 2 + len(LISTS_DATA) - 1,
                          "Lists & Dropdown Sources", level="title")
    col = 2
    for list_name, items in LISTS_DATA.items():
        ws.column_dimensions[get_column_letter(col)].width = 20
        hdr = ws.cell(row=3, column=col, value=list_name)
        hdr.font = FONT_COL_HDR
        hdr.fill = FILLS["column_header"]
        hdr.border = THIN_BORDER
        hdr.alignment = ALIGN_CENTER
        for j, item in enumerate(items):
            c = ws.cell(row=4 + j, column=col, value=item)
            c.font = FONT_BODY_11
            c.border = THIN_BORDER
        col += 1


def build_product_specs(ws):
    set_col_widths(ws, {"A": 3, "B": 20, "C": 14, "D": 14, "E": 14, "F": 14})
    apply_section_header(ws, 1, 2, 6, "Product Specifications", level="title")
    headers = ["Product Type", "AP (m²/m³)", "Wall Thickness (mm)",
               "Pitch (mm)", "Channel Type"]
    col_headers(ws, 3, 2, headers)
    for r in range(4, 9):
        for c in range(2, 7):
            cell = ws.cell(row=r, column=c)
            cell.font = FONT_BODY_11
            cell.border = THIN_BORDER
        ws.row_dimensions[r].height = RH_DATA


def build_backing_table(ws, table_def):
    table_name = table_def["table_name"]
    columns = table_def["columns"]
    for i in range(len(columns)):
        ws.column_dimensions[get_column_letter(i + 1)].width = 16
    create_excel_table(ws, table_name, columns, header_row=1, col_start=1, data_rows=1)


# ═══════════════════════════════════════════════════════════════════════════════
# Named ranges
# ═══════════════════════════════════════════════════════════════════════════════

def create_named_ranges(wb):
    # Control
    for i, (field, _) in enumerate(CONTROL_FIELDS):
        r = 3 + i
        dn = DefinedName(field, attr_text=f"'Control'!$C${r}")
        wb.defined_names.add(dn)

    # Specifications
    for i, (name, _, _) in enumerate(SPEC_FIELDS):
        r = 4 + i
        dn = DefinedName(name, attr_text=f"'Specifications'!$C${r}")
        wb.defined_names.add(dn)

    # Constants
    for i, (name, _, _) in enumerate(CONSTANTS):
        r = 4 + i
        dn = DefinedName(name, attr_text=f"'Constants'!$C${r}")
        wb.defined_names.add(dn)

    # GC_ ranges
    for i, name in enumerate(GC_NAMED_RANGES):
        r = 61 + i
        dn = DefinedName(name, attr_text=f"'Geometry Calc'!$C${r}")
        wb.defined_names.add(dn)

    # Lists
    col = 2
    for list_name, items in LISTS_DATA.items():
        cl = get_column_letter(col)
        end_row = 3 + len(items)
        dn = DefinedName(f"List_{list_name}", attr_text=f"'Lists'!${cl}$4:${cl}${end_row}")
        wb.defined_names.add(dn)
        col += 1


# ═══════════════════════════════════════════════════════════════════════════════
# Main build
# ═══════════════════════════════════════════════════════════════════════════════

def build():
    wb = Workbook()
    wb.remove(wb.active)

    sheets = {}
    for name, visibility in SHEET_DEFS:
        ws = wb.create_sheet(title=name)
        if visibility == "hidden":
            ws.sheet_state = "hidden"
        elif visibility == "veryHidden":
            ws.sheet_state = "veryHidden"
        sheets[name] = ws

    # ── Build each sheet ─────────────────────────────────────────────────
    build_home(sheets["Home"])
    build_specifications(sheets["Specifications"])
    build_hc_geometry(sheets["HC Geometry"])
    build_plate_geometry(sheets["Plate Geometry"])
    build_corrugated_geometry(sheets["Corrugated Geometry"])
    build_setup_summary(sheets["Setup Summary"])
    build_activity(sheets["Activity"])
    build_conversion(sheets["Conversion"])
    build_dp(sheets["DP"])
    build_geometry_calc(sheets["Geometry Calc"])
    build_control(sheets["Control"])
    build_constants(sheets["Constants"])
    build_lists(sheets["Lists"])
    build_product_specs(sheets["Product Specs"])

    for sheet_name, table_def in BACKING_TABLES.items():
        build_backing_table(sheets[sheet_name], table_def)

    create_named_ranges(wb)

    # ── Save ─────────────────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"Workbook saved to: {OUTPUT_FILE}")
    print(f"  Sheets: {len(wb.sheetnames)}")
    print(f"  Named ranges: {len(wb.defined_names)}")

    visible = [n for n, v in SHEET_DEFS if v == "visible"]
    hidden = [n for n, v in SHEET_DEFS if v == "hidden"]
    very_hidden = [n for n, v in SHEET_DEFS if v == "veryHidden"]
    print(f"\n  Visible ({len(visible)}):     {', '.join(visible)}")
    print(f"  Hidden ({len(hidden)}):      {', '.join(hidden)}")
    print(f"  VeryHidden ({len(very_hidden)}): {', '.join(very_hidden)}")

    tables = [t["table_name"] for t in BACKING_TABLES.values()]
    print(f"\n  Tables ({len(tables)}): {', '.join(tables)}")


if __name__ == "__main__":
    build()
